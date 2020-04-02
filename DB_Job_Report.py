def job_report_program():
    
    from openpyxl import load_workbook
    
    while True:
        try:
            
            file = str(input("Please enter the name of the job report: "))
        except ValueError:
            print("Sorry, I didn't understand that.")
            
            continue
        else:
            
            break



    file = file+".xlsx"

    wb = load_workbook(file)
    sheet = wb['Planning']
    
    while True:
        try:
            
            state = int(input("What do you want to do?\n [1] Update work hours\n [2] Change rate "))
        except ValueError:
            print("Sorry, I didn't understand that.")
            
            continue
        else:
        
            break
    if state == 1: 
        update_hours()
    elif state == 2:
        change_rate()

def update_hours():
    from openpyxl import load_workbook
    wb = load_workbook(file)
    sheet = wb['Planning']
    
    while True:
        try:
            
            column = int(input("Which workshift?\n [1] "+str(sheet.cell(row =5, column = 8).value)+"\n [2] "+str(sheet.cell(row =5, column = 9).value)+"\n [3] "+str(sheet.cell(row =5, column = 10).value)+"\n [4] "+str(sheet.cell(row =5, column = 11).value)+"\n [5] "+str(sheet.cell(row =5, column = 12).value)+"\n [6] "+str(sheet.cell(row =5, column = 13).value)+"\n [7] "+str(sheet.cell(row =5, column = 14).value)+"\n [8] "+str(sheet.cell(row =5, column = 15).value)+"\n [9] "+str(sheet.cell(row =5, column = 16).value)+"\n [10] "+str(sheet.cell(row =5, column = 17).value)+"\n [11] "+str(sheet.cell(row =5, column = 18).value)+"\n [12] "+str(sheet.cell(row =5, column = 19).value)+"\n [13] "+str(sheet.cell(row =5, column = 20).value)+"\n [14] "+str(sheet.cell(row =5, column = 21).value)+"\n [15] "+str(sheet.cell(row =5, column = 22).value)+"\n [16] "+str(sheet.cell(row =5, column = 23).value)+"\n [17] "+str(sheet.cell(row =5, column = 24).value)+"\n [18] "+str(sheet.cell(row =5, column = 25).value)+"\n [19] "+str(sheet.cell(row =5, column = 26).value)+"\n [20] "+str(sheet.cell(row =5, column = 27).value)+"\n [21] "+str(sheet.cell(row =5, column = 28).value)+"\n [22] "+str(sheet.cell(row =5, column = 29).value)+"\n [23] "+str(sheet.cell(row =5, column = 30).value)+"\n [24] "+str(sheet.cell(row =5, column = 31).value)+"\n [25] "+str(sheet.cell(row =5, column = 32).value)+"\n [26] "+str(sheet.cell(row =5, column = 33).value)+"\n [27] "+str(sheet.cell(row =5, column = 34).value)+"\n [28] "+str(sheet.cell(row =5, column = 35).value)+"\n [29] "+str(sheet.cell(row =5, column = 36).value)+"\n [30] "+str(sheet.cell(row =5, column = 37).value)+"\n [31] "+str(sheet.cell(row =5, column = 38).value)+"\n [32] "+str(sheet.cell(row =5, column = 39).value)+"\n [33] "+str(sheet.cell(row =5, column = 40).value)+"\n [34] "+str(sheet.cell(row =5, column = 41).value)+"\n [35] "+str(sheet.cell(row =5, column = 42).value)+"\n [36] "+str(sheet.cell(row =5, column = 43).value)+"\n [37] "+str(sheet.cell(row =5, column = 44).value)+"\n [38] "+str(sheet.cell(row =5, column = 45).value)+"\n [39] "+str(sheet.cell(row =5, column = 46).value)+"\n [40] "+str(sheet.cell(row =5, column = 47).value)+"\n [41] "+str(sheet.cell(row =5, column = 48).value)+"\n [42] "+str(sheet.cell(row =5, column = 49).value)+"\n [43] "+str(sheet.cell(row =5, column = 50).value)+"\n [44] "+str(sheet.cell(row =5, column = 51).value)+"\n [45] "+str(sheet.cell(row =5, column = 52).value)+"\n [46] "+str(sheet.cell(row =5, column = 53).value)+"\n [47] "+str(sheet.cell(row =5, column = 54).value)+"\n [48] "+str(sheet.cell(row =5, column = 55).value)+"\n: "))
        except ValueError:
            print("Sorry, I didn't understand that.")
            
            continue
        else:
            
            break
    if column == 1: 
        column = 8
    elif column == 2:
        column = 9
    elif column == 3:
        column = 10
    elif column == 4:
        column = 11
    elif column == 5:
        column = 12
    elif column == 6:
        column = 13
    elif column == 7:
        column = 14
    elif column == 8:
        column = 15
    elif column == 9:
        column = 16
    elif column == 10:
        column = 17
    elif column == 11:
        column = 18
    elif column == 12:
        column = 19
    elif column == 13:
        column = 20
    elif column == 14:
        column = 21
    elif column == 15:
        column = 22
    elif column == 16:
        column = 23
    elif column == 17:
        column = 24
    elif column == 18:
        column = 25
    elif column == 19:
        column = 26
    elif column == 20:
        column = 27
    elif column == 21:
        column = 28
    elif column == 22:
        column = 29
    elif column == 23:
        column = 30
    elif column == 24:
        column = 31
    elif column == 25:
        column = 32
    elif column == 26:
        column = 33
    elif column == 27:
        column = 34
    elif column == 28:
        column = 35
    elif column == 29:
        column = 36
    elif column == 30:
        column = 37
    elif column == 31:
        column = 38
    elif column == 32:
        column = 39
    elif column == 33:
        column = 40
    elif column == 34:
        column = 41
    elif column == 35:
        column = 42
    elif column == 36:
        column = 43
    elif column == 37:
        column = 44
    elif column == 38:
        column = 45
    elif column == 39:
        column = 46
    elif column == 40:
        column = 47
    elif column == 41:
        column = 48
    elif column == 42:
        column = 49
    elif column == 43:
        column = 50
    elif column == 44:
        column = 51
    elif column == 45:
        column = 52
    elif column == 46:
        column = 53
    elif column == 47:
        column = 54
    elif column == 48:
        column = 55
    else:
        print("Not a viable answer, try again.")
    
    while True:
        try:
            
            row = int(input("Which week?\n [1] "+str(sheet.cell(row =7, column = 7).value)+"\n [2] "+str(sheet.cell(row =8, column = 7).value)+"\n [3] "+str(sheet.cell(row =7, column = 8).value)+"\n [4] "+str(sheet.cell(row =9, column = 7).value)+"\n [5] "+str(sheet.cell(row =10, column = 7).value)+"\n [6] "+str(sheet.cell(row =11, column = 7).value)+"\n [7] "+str(sheet.cell(row =12, column = 7).value)+"\n [8] "+str(sheet.cell(row =13, column = 7).value)+"\n [9] "+str(sheet.cell(row =14, column = 7).value)+"\n [10] "+str(sheet.cell(row =15, column = 7).value)+"\n: "))
        except ValueError:
            print("Sorry, I didn't understand that.")
            
            continue
        else:
            
            break   
    if row == 1:
        row = 7
    elif row == 2:
        row = 8
    elif row == 3:
        row = 9
    elif row == 4:
        row = 10
    elif row == 5:
        row = 11
    elif row == 6:
        row = 12
    elif row == 7:
        row = 13
    elif row == 8:
        row = 14
    elif row == 9:
        row = 15
    elif row == 10:
        row = 16
    else:
        print("Not a viable answer, try again.")
    
    while True:
        try:
            
            value = int(input("Please enter the number of hours: "))
        except ValueError:
            print("Sorry, I didn't understand that.")
            
            continue
        else:
            
            break

    a = sheet.cell(row=row,column=column).value
    ff = sheet.cell(row =1000, column = 1000).value


    if type(ff) == type(a):
        sheet.cell(row=row,column=column,value=0)
        wb.save(file)
        sheet.cell(row=row,column=column,value=int(a)+value)
        wb.save(file)
    else:
        b = sheet.cell(row=row,column=column,value=int(a)+value)
        b
        wb.save(file)

    print("Completed")

def change_rate():
    from openpyxl import load_workbook
    wb = load_workbook(file)
    sheet = wb['Planning']
    
    while True:
        try:
            
            column = int(input("Which workshift?\n [1] "+str(sheet.cell(row =5, column = 8).value)+"\n [2] "+str(sheet.cell(row =5, column = 9).value)+"\n [3] "+str(sheet.cell(row =5, column = 10).value)+"\n [4] "+str(sheet.cell(row =5, column = 11).value)+"\n [5] "+str(sheet.cell(row =5, column = 12).value)+"\n [6] "+str(sheet.cell(row =5, column = 13).value)+"\n [7] "+str(sheet.cell(row =5, column = 14).value)+"\n [8] "+str(sheet.cell(row =5, column = 15).value)+"\n [9] "+str(sheet.cell(row =5, column = 16).value)+"\n [10] "+str(sheet.cell(row =5, column = 17).value)+"\n [11] "+str(sheet.cell(row =5, column = 18).value)+"\n [12] "+str(sheet.cell(row =5, column = 19).value)+"\n [13] "+str(sheet.cell(row =5, column = 20).value)+"\n [14] "+str(sheet.cell(row =5, column = 21).value)+"\n [15] "+str(sheet.cell(row =5, column = 22).value)+"\n [16] "+str(sheet.cell(row =5, column = 23).value)+"\n [17] "+str(sheet.cell(row =5, column = 24).value)+"\n [18] "+str(sheet.cell(row =5, column = 25).value)+"\n [19] "+str(sheet.cell(row =5, column = 26).value)+"\n [20] "+str(sheet.cell(row =5, column = 27).value)+"\n [21] "+str(sheet.cell(row =5, column = 28).value)+"\n [22] "+str(sheet.cell(row =5, column = 29).value)+"\n [23] "+str(sheet.cell(row =5, column = 30).value)+"\n [24] "+str(sheet.cell(row =5, column = 31).value)+"\n [25] "+str(sheet.cell(row =5, column = 32).value)+"\n [26] "+str(sheet.cell(row =5, column = 33).value)+"\n [27] "+str(sheet.cell(row =5, column = 34).value)+"\n [28] "+str(sheet.cell(row =5, column = 35).value)+"\n [29] "+str(sheet.cell(row =5, column = 36).value)+"\n [30] "+str(sheet.cell(row =5, column = 37).value)+"\n [31] "+str(sheet.cell(row =5, column = 38).value)+"\n [32] "+str(sheet.cell(row =5, column = 39).value)+"\n [33] "+str(sheet.cell(row =5, column = 40).value)+"\n [34] "+str(sheet.cell(row =5, column = 41).value)+"\n [35] "+str(sheet.cell(row =5, column = 42).value)+"\n [36] "+str(sheet.cell(row =5, column = 43).value)+"\n [37] "+str(sheet.cell(row =5, column = 44).value)+"\n [38] "+str(sheet.cell(row =5, column = 45).value)+"\n [39] "+str(sheet.cell(row =5, column = 46).value)+"\n [40] "+str(sheet.cell(row =5, column = 47).value)+"\n [41] "+str(sheet.cell(row =5, column = 48).value)+"\n [42] "+str(sheet.cell(row =5, column = 49).value)+"\n [43] "+str(sheet.cell(row =5, column = 50).value)+"\n [44] "+str(sheet.cell(row =5, column = 51).value)+"\n [45] "+str(sheet.cell(row =5, column = 52).value)+"\n [46] "+str(sheet.cell(row =5, column = 53).value)+"\n [47] "+str(sheet.cell(row =5, column = 54).value)+"\n [48] "+str(sheet.cell(row =5, column = 55).value)+"\n: "))
        except ValueError:
            print("Sorry, I didn't understand that.")
            
            continue
        else:
            
            break
    if column == 1: 
        column = 8
    elif column == 2:
        column = 9
    elif column == 3:
        column = 10
    elif column == 4:
        column = 11
    elif column == 5:
        column = 12
    elif column == 6:
        column = 13
    elif column == 7:
        column = 14
    elif column == 8:
        column = 15
    elif column == 9:
        column = 16
    elif column == 10:
        column = 17
    elif column == 11:
        column = 18
    elif column == 12:
        column = 19
    elif column == 13:
        column = 20
    elif column == 14:
        column = 21
    elif column == 15:
        column = 22
    elif column == 16:
        column = 23
    elif column == 17:
        column = 24
    elif column == 18:
        column = 25
    elif column == 19:
        column = 26
    elif column == 20:
        column = 27
    elif column == 21:
        column = 28
    elif column == 22:
        column = 29
    elif column == 23:
        column = 30
    elif column == 24:
        column = 31
    elif column == 25:
        column = 32
    elif column == 26:
        column = 33
    elif column == 27:
        column = 34
    elif column == 28:
        column = 35
    elif column == 29:
        column = 36
    elif column == 30:
        column = 37
    elif column == 31:
        column = 38
    elif column == 32:
        column = 39
    elif column == 33:
        column = 40
    elif column == 34:
        column = 41
    elif column == 35:
        column = 42
    elif column == 36:
        column = 43
    elif column == 37:
        column = 44
    elif column == 38:
        column = 45
    elif column == 39:
        column = 46
    elif column == 40:
        column = 47
    elif column == 41:
        column = 48
    elif column == 42:
        column = 49
    elif column == 43:
        column = 50
    elif column == 44:
        column = 51
    elif column == 45:
        column = 52
    elif column == 46:
        column = 53
    elif column == 47:
        column = 54
    elif column == 48:
        column = 55
    else:
        print("Not a viable answer, try again.")
    
    while True:
        try:
            
            value = int(input("Please enter the new rate: "))
        except ValueError:
            print("Sorry, I didn't understand that.")
            
            continue
        else:
            
            break
        
    sheet.cell(row=44, column=column, value=value)
    wb.save(file)

    print("Completed")
 


